import csv
from pathlib import Path
from typing import Dict, Tuple

from docx import Document
from openpyxl import load_workbook
from PyPDF2 import PdfReader


class DocumentParser:
    """Read supported file types and return extracted text + metadata."""

    SUPPORTED = {".pdf", ".docx", ".xlsx", ".csv", ".txt"}

    def parse(self, file_path: Path) -> Tuple[str, Dict[str, str]]:
        suffix = file_path.suffix.lower()
        if suffix not in self.SUPPORTED:
            raise ValueError(f"Unsupported file type: {suffix}")

        if suffix == ".pdf":
            text = self._parse_pdf(file_path)
        elif suffix == ".docx":
            text = self._parse_docx(file_path)
        elif suffix == ".xlsx":
            text = self._parse_xlsx(file_path)
        elif suffix == ".csv":
            text = self._parse_csv(file_path)
        else:
            text = file_path.read_text(encoding="utf-8", errors="ignore")

        metadata = {
            "filename": file_path.name,
            "path": str(file_path),
            "extension": suffix,
        }
        return text, metadata

    def _parse_pdf(self, file_path: Path) -> str:
        reader = PdfReader(str(file_path))
        pages = [page.extract_text() or "" for page in reader.pages]
        return "\n".join(pages)

    def _parse_docx(self, file_path: Path) -> str:
        doc = Document(str(file_path))
        return "\n".join(paragraph.text for paragraph in doc.paragraphs)

    def _parse_xlsx(self, file_path: Path) -> str:
        workbook = load_workbook(filename=str(file_path), data_only=True)
        chunks = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append(" | ".join("" if cell is None else str(cell) for cell in row))
            chunks.append(f"Sheet: {sheet_name}\n" + "\n".join(rows))
        return "\n\n".join(chunks)

    def _parse_csv(self, file_path: Path) -> str:
        with file_path.open(newline="", encoding="utf-8", errors="ignore") as handle:
            reader = csv.reader(handle)
            return "\n".join(" | ".join(cell for cell in row) for row in reader)
